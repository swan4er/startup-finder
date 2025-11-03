#!/usr/bin/env python3
# -*- coding: utf-8 -*-

import ssl
ssl._create_default_https_context = ssl._create_unverified_context

import sys
import os
from openpyxl import Workbook, load_workbook
from openpyxl.styles import Font
from datetime import datetime

from config_manager import get_producthunt_token
from producthunt_parser import ProductHuntParser
from utils import resolve_urls_batch
from crunchbase_parser import CrunchbaseParser


def get_user_input():
    """Получает параметры парсинга от пользователя"""
    print("\n" + "="*60)
    print("НАСТРОЙКИ ПАРСИНГА")
    print("="*60)
    
    # Годы
    years_input = input("За сколько лет собирать проекты? [По умолчанию: 3]: ").strip() or "3"
    years = int(years_input)
    
    # Черный список
    blacklist_input = input("Черный список слов через запятую [По умолчанию: пусто]: ").strip()
    blacklist = [word.strip() for word in blacklist_input.split(',') if word.strip()]
    
    # Максимум сотрудников
    max_makers_input = input("Максимальное количество сотрудников [По умолчанию: 10]: ").strip() or "10"
    max_makers = int(max_makers_input)
    
    # Лимит проектов
    max_products_input = input("Лимит количества проектов [По умолчанию: 5000]: ").strip() or "5000"
    max_products = int(max_products_input)
    
    print("="*60 + "\n")
    
    return {
        'years': years,
        'blacklist': blacklist,
        'max_makers': max_makers,
        'max_products': max_products
    }


def save_to_excel(products, filename='producthunt.xlsx', include_crunchbase=False):
    """Сохраняет продукты в Excel файл"""
    print(f"\n💾 Сохранение результатов в {filename}...")
    
    wb = Workbook()
    ws = wb.active
    ws.title = "Products"
    
    # Заголовки
    headers = ['name', 'description', 'votesCount', 'website', 'producthunt_url', 'makers', 'created_at']
    
    if include_crunchbase:
        headers.extend(['crunchbase_url', 'funding_amount'])
    
    ws.append(headers)
    
    # Стилизация заголовков
    for cell in ws[1]:
        cell.font = Font(bold=True)
    
    # Данные
    for product in products:
        row = [
            product.get('name', ''),
            product.get('description', ''),
            product.get('votesCount', 0),
            product.get('website', ''),
            product.get('producthunt_url', ''),
            product.get('makers', 0),
            product.get('created_at', '')
        ]
        
        if include_crunchbase:
            row.extend([
                product.get('crunchbase_url', ''),
                product.get('funding_amount', '')
            ])
        
        ws.append(row)
    
    # Автоширина колонок
    for column in ws.columns:
        max_length = 0
        column_letter = column[0].column_letter
        for cell in column:
            try:
                if len(str(cell.value)) > max_length:
                    max_length = len(str(cell.value))
            except:
                pass
        adjusted_width = min(max_length + 2, 50)
        ws.column_dimensions[column_letter].width = adjusted_width
    
    wb.save(filename)
    print(f"✓ Файл сохранен: {filename}")
    print(f"  Записей: {len(products)}")


def load_products_from_excel(filename='producthunt.xlsx'):
    """Загружает продукты из существующего Excel файла"""
    try:
        wb = load_workbook(filename)
        ws = wb.active
        
        products = []
        headers = [cell.value for cell in ws[1]]
        
        for row in ws.iter_rows(min_row=2, values_only=True):
            product = {}
            for i, value in enumerate(row):
                if i < len(headers):
                    product[headers[i]] = value
            products.append(product)
        
        return products
    except Exception as e:
        print(f"❌ Ошибка загрузки файла: {e}")
        return None


def check_existing_table():
    """Проверяет наличие существующей таблицы и предлагает варианты"""
    filename = 'producthunt.xlsx'
    
    if os.path.exists(filename):
        print("\n" + "="*60)
        print("📋 НАЙДЕНА СУЩЕСТВУЮЩАЯ ТАБЛИЦА")
        print("="*60)
        print(f"Файл: {filename}")
        
        # Получаем информацию о файле
        try:
            wb = load_workbook(filename)
            ws = wb.active
            row_count = ws.max_row - 1  # минус заголовок
            print(f"Проектов в таблице: {row_count}")
        except:
            print("Не удалось прочитать информацию о файле")
        
        print("\nВыберите действие:")
        print("1. Начать парсинг ProductHunt с нуля (старая таблица будет перезаписана)")
        print("2. Продолжить с Crunchbase парсингом (использовать существующую таблицу)")
        print("3. Выход")
        print("="*60)
        
        choice = input("\nВаш выбор [По умолчанию: 1]: ").strip()
        
        if choice == '2':
            return 'crunchbase'
        elif choice == '3':
            print("\n👋 До свидания!")
            return 'exit'
        else:
            # По умолчанию или если выбрано 1
            confirm = input("\n⚠ Старая таблица будет удалена. Продолжить? (y/n) [По умолчанию: y]: ").strip().lower()
            if confirm in ['n', 'no', 'н', 'нет']:
                print("\n👋 До свидания!")
                return 'exit'
            return 'new'
    
    return 'new'


def main():
    try:
        print("\n" + "="*60)
        print("🚀 ПОИСК ИДЕИ ДЛЯ СТАРТАПА")
        print("="*60)
        
        # Проверяем наличие существующей таблицы
        mode = check_existing_table()
        
        if mode == 'exit':
            return
        
        if mode == 'crunchbase':
            # Загружаем данные из существующего файла
            print("\n📂 Загрузка данных из producthunt.xlsx...")
            products = load_products_from_excel()
            
            if not products:
                print("❌ Не удалось загрузить данные из файла")
                return
            
            print(f"✓ Загружено проектов: {len(products)}")
            
            # Сразу переходим к Crunchbase
            print("\n" + "="*60)
            
            # Создаем парсер Crunchbase
            crunchbase = CrunchbaseParser()
            
            # Авторизация на Crunchbase
            crunchbase.setup_authentication()
            
            # Парсинг Crunchbase - поиск компаний
            products = crunchbase.search_organizations_batch(products)
            
            # Парсинг Crunchbase - получение funding
            products = crunchbase.get_funding_amounts_batch(products)
            
            # Шаг 9: Сохранение финального результата
            save_to_excel(products, include_crunchbase=True)
            
            print("\n" + "="*60)
            print("✅ ПАРСИНГ CRUNCHBASE ЗАВЕРШЕН")
            print("="*60)
            print(f"Итоговое количество проектов: {len(products)}")
            
            # Статистика по Crunchbase
            cb_found = sum(1 for p in products if p.get('crunchbase_url'))
            funding_found = sum(1 for p in products if p.get('funding_amount'))
            
            print(f"Найдено на Crunchbase: {cb_found}")
            print(f"С данными о финансировании: {funding_found}")
            print(f"Файл: producthunt.xlsx")
            print("="*60 + "\n")
            
            return
        
        # Режим 'new' - начинаем с нуля
        # Шаг 1: Получить токен ProductHunt
        token = get_producthunt_token()
        
        # Шаг 2: Получить параметры парсинга
        params = get_user_input()
        
        # Шаг 3: Парсинг ProductHunt
        parser = ProductHuntParser(
            token=token,
            years=params['years'],
            blacklist=params['blacklist'],
            max_makers=params['max_makers'],
            max_products=params['max_products']
        )
        
        products = parser.parse()
        
        if not products:
            print("\n❌ Не найдено ни одного продукта по заданным критериям")
            return
        
        # Шаг 4: Резолв URL и проверка доступности
        products = resolve_urls_batch(products, max_workers=20)
        
        if not products:
            print("\n❌ Все сайты недоступны")
            return
        
        # Вопрос о продолжении
        print("\n" + "="*60)
        continue_input = input("Продолжить с парсингом Crunchbase? (Y/n): ").strip().lower()

        save_to_excel(products, include_crunchbase=False)
        
        if continue_input not in ['', 'y', 'yes', 'д', 'да']:
            print("\n" + "="*60)
            print("✅ ПАРСИНГ PRODUCTHUNT ЗАВЕРШЕН")
            print("="*60)
            print(f"Итоговое количество проектов: {len(products)}")
            print(f"Файл: producthunt.xlsx")
            print("="*60 + "\n")
            return
        
        # Создаем парсер Crunchbase
        crunchbase = CrunchbaseParser()
        
        # Авторизация на Crunchbase
        crunchbase.setup_authentication()
        
        # Парсинг Crunchbase - поиск компаний
        products = crunchbase.search_organizations_batch(products)
        
        # Парсинг Crunchbase - получение funding
        products = crunchbase.get_funding_amounts_batch(products)
        
        # Шаг 9: Сохранение финального результата
        save_to_excel(products, include_crunchbase=True)
        
        print("\n" + "="*60)
        print("✅ ПАРСИНГ ПОЛНОСТЬЮ ЗАВЕРШЕН")
        print("="*60)
        print(f"Итоговое количество проектов: {len(products)}")
        
        # Статистика по Crunchbase
        cb_found = sum(1 for p in products if p.get('crunchbase_url'))
        funding_found = sum(1 for p in products if p.get('funding_amount'))
        
        print(f"Найдено на Crunchbase: {cb_found}")
        print(f"С данными о финансировании: {funding_found}")
        print(f"Файл: producthunt.xlsx")
        print("="*60 + "\n")
        
    except KeyboardInterrupt:
        print("\n\n⚠ Прервано пользователем")
        sys.exit(0)
    except Exception as e:
        print(f"\n❌ Ошибка: {e}")
        import traceback
        traceback.print_exc()
        sys.exit(1)


if __name__ == "__main__":
    main()

